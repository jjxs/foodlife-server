from os import walk
from django.http import HttpResponse
from django.conf import settings
from app.log import logger
import openpyxl as px
import sys
import os
import csv

class ExcelFileStord(object):
    __store = {}

    @classmethod
    def add(cls, key, excel):
        cls.__store[key] = excel

    @classmethod
    def get(cls, id):

        if cls.__store is None or len(cls.__store) == 0:
            return None

        if id in cls.__store:
            return cls.__store[id]

        return None

def init_excel_cache():

    try:
        path = settings.EXCEL_TEMPLATE_PATH

        if not os.path.exists(path):
            logger.error("パス「{0}」は存在しない".format(path))
            return

        for (dirpath, dirnames, filenames) in walk(path):
            for filename in filenames:
                if not filename.endswith(".xlsx"):
                    continue

                dir_path = dirpath
                filepath = os.path.join(dirpath, filename)
                tmp = dir_path.replace('/', '\\').split('\\')[-1] + '.' + filename.replace('.xlsx', '')
                ExcelFileStord.add(tmp, filepath)

    except Exception:
        logger.error("Unexpected error: {0} \n ※Excelファイル問題があるか、ご確かめてください！".format(sys.exc_info()[0]))


def get_excel_info(self, fileId, data):
    '''Excelファイル内容読み込む'''
    '''# 上書きデータ：フォーマット：{シート名:[{項目1：xxxx,項目2：xxxx},{項目1：xxxx,項目2：xxxx}]}'''

    strFilePath = ExcelFileStord.get(fileId)

    if not strFilePath:
        init_excel_cache()
        strFilePath = ExcelFileStord.get(fileId)

    if not strFilePath:
        logger.error("該当Excelテンプレート存在しません。{0}".format(fileId))        
        return

    response = HttpResponse(content_type='application/vnd.ms-excel')

    wb = px.load_workbook(strFilePath)

    sheetNames = wb.get_sheet_names()

    for sheetName in sheetNames:
        get_excel_info_by_sheet(self, wb, sheetName, data)

    # 不要なシートを削除する
    remove_sheet(self, wb, sheetNames, data.keys())

    wb.save(response)

    return response

def get_excel_info_by_sheet(self, wb, sheetName, data):
    '''各シートの情報を取得し、値を上書き'''

    sh = wb[sheetName]
    # 該当シート対するデータが存在していない場合、'No Data'を返却
    tmpData = data.get(sheetName, 'No Data')

    if tmpData == 'No Data':
        return

    maxrow = sh.max_row
    maxcol = sh.max_column

    for row in range(maxrow):
        for col in range(maxcol):
            # セル取得
            cell = sh.cell(row + 1, col + 1)
            # 該当セルの値
            tmp = str(cell.value)
            
            # 上書きの場合
            if tmp.startswith(':'):
                # 該当シートの内容がデータに存在していない場合、空白を設定
                if tmpData == 'No Data':
                    cell.value = ''
                elif len(tmpData) == 0:
                    cell.value = ''
                else:
                    curRow = row
                    curCol = col
                    for dic in tmpData:
                        tmpValue = dic.get(tmp.replace(':', ''), '')

                        if not tmpValue:
                            tmpValue = '';

                        sh.cell(curRow + 1, curCol + 1, tmpValue)
                        curRow += 1

def remove_sheet(self, wb, sheet_names, data_keys):
    '''不要なシートを削除'''
    sheet_index = 0
    for sheet_name in sheet_names:
        if not sheet_name in data_keys:
            wb.remove(sheet_index)
        sheet_index = sheet_index + 1

def download_csv(self, header, data, excludeItem=[]):
    '''CSVファイルダウンロード'''

    response = HttpResponse(content_type='text/csv; charset=shift-jis')

    writer = csv.writer(response)

    writer.writerow(header)

    datarows = []
    for obj in data:
        tmp = []
        for key, value in obj.items():
            if (not excludeItem or key not in excludeItem):
                if not value:
                    value = ''
                tmp.append(str(value).encode('shift-jis','ignore').decode('shift-jis','ignore'))

        datarows.append(tmp)

    writer.writerows(datarows)

    return response